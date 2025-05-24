import requests
import sqlite3
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from datetime import datetime

dados_api = []
url = 'https://books.toscrape.com/'
dados = []

#gerarRelatorioPaises

def gerarRelatorioPaises(nome, nome_ofc, capital, continente, regiao, sub_reg, populacao, area, idioma, moeda_nome, moeda_simb, fuso, url_bandeira):
    conexao = sqlite3.connect("paises.db") #Abrimos ou criamos um arquivo com extensão ao BD chamado paises. Também criamos um obj chamado extensão
    cursor = conexao.cursor() # Criamos um cursor, um objeto intermediario para executar comandos SQL dentro do BD. Com ele podemos fazer CREAT TABLE, INSERT INT, SELECT

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS paises(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome TEXT,
                nome_oficial TEXT,
                capital TEXT,
                continente TEXT,
                regiao TEXT,
                sub_regiao TEXT,
                populacao INTEGER,
                area REAL,
                idioma TEXT,
                moeda_nome TEXT,
                moeda_simbolo TEXT,
                fuso TEXT,
                url_bandeira TEXT
            )
        ''')
    
    cursor.execute('''
        INSERT INTO paises(
                   nome, nome_oficial, capital, continente, regiao, sub_regiao,
                   populacao, area, idioma, moeda_nome, moeda_simbolo,fuso, url_bandeira
                   ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
                ''', (
                    nome, nome_ofc, capital, continente, regiao, sub_reg,
                    populacao, area, idioma, moeda_nome, moeda_simb, fuso, url_bandeira
                ))
    conexao.commit() #Estamos deixando registrado nossa ação no banco de dados
    conexao.close()

    dados_api.append({
        'Nome': nome,
        'Nome Oficial': nome_ofc,
        'Capital': capital,
        'Continente': continente,
        'Região': regiao,
        'Sub-Região': sub_reg,
        'População': populacao,
        'Área': area,
        'Moeda': moeda_nome,
        "Símbolo da Moeda": moeda_simb,
        'Idioma': idioma,
        'Fuso Horário': fuso,
        'URL da Bandeira': url_bandeira
    })
    # df = pd.DataFrame(dados_api)

    # df.to_excel('paisesRelatorio.xlsx', index=False, engine='openpyxl')
    # wb = load_workbook('relatorioProjeto.xlsx')
    # ws = wb.active

    # for call in ws[1]:
    #     call.font = Font(bold=True)
    # wb.save('paises_formatados')
    # print(f"Dados do país {nome} armazenados.")
    print(f"Dados do país {nome} armazenados.")


def gerarRelatorioLivros(titulo, preco, disponibilidade, estrela):
    conexao = sqlite3.connect('livraria.db')
    cursor = conexao.cursor()

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS livros(
                   id INTEGER PRIMARY KEY AUTOINCREMENT,
                   titulo TEXT,
                   preco TEXT,
                   disponibilidade TEXT,
                   estrela INTEGER
                   )
            ''')
    cursor.execute('''
        INSERT INTO livros(
                   titulo, preco,
                   disponibilidade, estrela
                   ) VALUES (?,?,?,?)
                   ''',(
                       titulo, preco, disponibilidade, estrela
                   ))
    conexao.commit()
    conexao.close()

    dados.append({
        'Título':titulo,
        'Preço': preco,
        'Disponibilidade': disponibilidade,
        'Estrelas': estrela
    })


    df_livros = pd.DataFrame(dados)

    #%Y = ano com 4 dígitos. %m = mês com dois dígitos...
    #agora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    #df_info = pd.DataFrame([["Este relatório foi gerado ", agora]], columns=["Informação", "Data"])

    #df_final = pd.concat([df_info, pd.DataFrame([[]]), df_livros], ignore_index=True)

    # Junção
    #df_final.to_excel('livrosRelatorio.xlsx', index=False, engine='openpyxl')
    
    # df = pd.DataFrame(dados_api)
    # df.to_excel('relatorioProjeto.xlsx', index=False, engine='openpyxl')

def extrairDadosLivros(url_base):
    resposta = requests.get(url_base)
    soup = BeautifulSoup(resposta.text, 'html.parser')

    livros = soup.find_all('article', class_='product_pod')
    dados = []
    for livro in livros[:10]:
        titulo = livro.h3.a['title']
        preco = livro.find('p', class_='price_color').text
        disponibilidade = livro.find('p', class_='instock availability').text.strip()

        estrela_tag = livro.find('p', class_='star-rating')
        estrela_text = estrela_tag.get('class')[1]

        mapa_estrela = {
            'One': 1,
            'Two': 2,
            'Three': 3,
            'Four': 4,
            'Five': 5
        }
        estrelas = mapa_estrela.get(estrela_text, 0)

        gerarRelatorioLivros(titulo, preco, disponibilidade, estrelas)


def extrairDadosPaises(pais):
    url = f"https://restcountries.com/v3.1/name/{pais}"
    resposta = requests.get(url)

    if resposta.status_code!=200:
        print({"requisição inválida":"Erro ao buscar os dados"})
        return
    
    dados = resposta.json()
    if not dados:
        print({f"Requisição inválida":"Erro ao buscar nome do país{pais}"})
        return
    
    pais = dados[0]

    try:
        nome = pais['name']['common'] # esta contido dentro de um dicionário, sendo assim acessamos pela chave que retorna o valor
        nome_ofc = pais['name']['official']
        capital = pais.get('capital', ['Sem capital'])[0] # Só necessário caso o campo esteja vazio. E capital nos retorna uma lista, por essa razão o zero
        # "capital": ["Brasília"]
        continente = pais.get('continents',['Desconhecido'])[0]
        regiao =  pais['region']
        sub_reg = pais['subregion']
        populacao = pais['population'] 
        area = pais['area']
        lingua = pais.get('languages',{}) # Pegamos o dicionário ou lista completa com a chave languages
        #Aqui está tranformando em lista o valor de languages e pegando exatamente o índice 0 que representa o idioma principal,
        #depois checa se a variável está vazia, se sim ela passa a valer sem idioma
        idioma = list(lingua.values())[0] if lingua else 'Sem idioma' 
        fuso = pais['timezones'][0]
        url_bandeira = pais.get('flags',{}).get('png','Sem URL') # Entramos no dicionário com nome flag me pais, caso não exista devolvemos {}. depois acessamos a chave png

        currencies = pais.get('currencies',{}) # Estamos pegando o dicionário completo de currencies, caso não exista o deixamos vazio.
        if currencies:
            moeda_info = list(currencies.values())[0] # Estamos acessando as informações de moesda, tranformando em lista por assim será fácil pegar o indice 0 que representa a moeda principa
            moeda_nome = moeda_info.get('name', 'Desconhecido') # Dentro de moeda_info pegamos o nome dela
            moeda_simb = moeda_info.get('symbol', 'N/A')

        

        gerarRelatorioPaises(nome, nome_ofc, capital, continente, regiao,
                        sub_reg, populacao, area, idioma, moeda_nome,
                        moeda_simb, fuso, url_bandeira)
    except Exception as e:
        print({"Erro ao processar os dados":str(e)})

def gerar_relatorio_geral():
    df_livros = pd.DataFrame(dados)
    df_paises = pd.DataFrame(dados_api)

    agora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    df_info = pd.DataFrame([["Este relatório foi gerado em", agora]], columns=["Descrição", "Data e Hora"])

    with pd.ExcelWriter("Relatorio_Geral.xlsx", engine="openpyxl") as writer:
        df_info.to_excel(writer, sheet_name="Informações", index=False)
        df_livros.to_excel(writer, sheet_name="Livros", index=False)
        df_paises.to_excel(writer, sheet_name="Países", index=False)

        print("Relatório geral gerado com sucesso: 'Relatorio_Geral.xlsx'")



extrairDadosLivros(url)

visualizar = sqlite3.connect('livraria.db')
cursor = visualizar.cursor()

print("Visualizando tabelas existentes: ")
cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
print(cursor.fetchall())

print("Visualizar dados da tabela: ")
cursor.execute("SELECT * FROM livros")
for linha in cursor.fetchall():
    print(linha)

visualizar.close()

pais = input("Digite o nome do País em Inglês: ")
extrairDadosPaises(pais)

for i in range(2):
    pais = input("Digite o nome de outro País, também em Inglês: ")
    extrairDadosPaises(pais)

visualizar = sqlite3.connect('paises.db')
cursor = visualizar.cursor()

print("Tabelas existentes: ")
cursor.execute('SELECT * FROM  paises')
for linha in cursor.fetchall():
    print(linha)


print("Dados existentes nas tabelas: ")
cursor.execute("SELECT name FROM sqlite_master WHERE type ='table';")
print(cursor.fetchall())

visualizar.close()

gerar_relatorio_geral()
print("Relatório geral realizado com êxito")